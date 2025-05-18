import os

from mospy import Account

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def generate(
        hrp: str,
        amount: int
) -> list:
    result = []
    for _ in range(amount):
        account: Account = Account(hrp=hrp)
        piece = {
            "address": account.address,
            "mnemonic": account.seed_phrase,
            "pk": account.private_key.hex()
        }

        if not piece["pk"].startswith("0x"):
            piece["pk"] = "0x" + piece["pk"]

        result.append(piece)

    return result


def main():
    hrp = input("1️⃣ Укажите заголовок, который будет использован при генерации кошельков "
                "(примеры – `osmo`, `cosmos` и т.д.): ")

    amount = -1
    while amount == -1:
        try:
            a = int(input("2️⃣ Укажите желаемое кол-во кошельков: "))
        except ValueError:
            print("⁉️ Неверный формат! Повторите попытку ввода.\n")
            continue

        if a <= 0:
            print("☝️ Кол-во должно быть больше 0! Повторите попытку ввода.\n")
        else:
            amount = a

    try:
        os.mkdir(os.getcwd() + os.sep + "output")
    except FileExistsError:
        pass

    filename = ""
    while filename == "":
        name = input("Укажите имя .xlsx-файла без расширения, куда будут записаны результаты: ")

        path = os.getcwd() + os.sep + "output" + os.sep + f"{name}.xlsx"
        if os.path.exists(path):
            print("💥 Файл уже существует! Укажите уникальное имя.")
            continue

        filename = path

    wallets = generate(hrp, amount)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cosmos"

    A1 = sheet.cell(row=1, column=1)
    A1.value = "Адрес"

    B1 = sheet.cell(row=1, column=2)
    B1.value = "Приватный ключ"

    C1 = sheet.cell(row=1, column=3)
    C1.value = "Мнемоника"

    C1.font = B1.font = A1.font = Font(
        name="Arial",
        bold=True,
        size=16,
        color="FFFFFF"
    )

    C1.fill = B1.fill = A1.fill = PatternFill(
        fill_type="solid",
        start_color="000000",
        end_color="000000"
    )

    C1.alignment = B1.alignment = A1.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    row = 2
    address_width_set = False
    pk_width_set = False
    mnemonic_width_set = False
    for wallet in wallets:
        address_cell = sheet.cell(row=row, column=1)
        address_cell.value = wallet["address"]
        address_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        address_cell.font = Font(
            name="Consolas",
            bold=True,
            size=12
        )

        if not address_width_set:
            letter = address_cell.column_letter
            sheet.column_dimensions[letter].width = len(address_cell.value) * 1.5
            address_width_set = True

        pk_cell = sheet.cell(row=row, column=2)
        pk_cell.value = wallet["pk"]
        pk_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        pk_cell.font = Font(
            name="Consolas",
            bold=False,
            size=12
        )

        if not pk_width_set:
            letter = pk_cell.column_letter
            sheet.column_dimensions[letter].width = len(pk_cell.value) * 1.25
            pk_width_set = True

        mnemonic_cell = sheet.cell(row=row, column=3)
        mnemonic_cell.value = wallet["mnemonic"]
        mnemonic_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        mnemonic_cell.font = Font(
            name="Consolas",
            bold=False,
            size=12
        )

        if not mnemonic_width_set:
            letter = mnemonic_cell.column_letter
            sheet.column_dimensions[letter].width = len(mnemonic_cell.value) * 1.5
            mnemonic_width_set = True

        row += 1

    workbook.save(filename)
    print("✅ Таблица успешно сохранена!")


if __name__ == "__main__":
    main()
